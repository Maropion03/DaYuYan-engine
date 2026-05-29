import base64
import io
import json
import mimetypes
import re
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional

import fitz
import pytesseract
import requests
from PIL import Image

from scene_runtime.contracts import SceneRequest


ROOT_DIR = Path(__file__).resolve().parent.parent
TESSERACT_CMD = Path(r"C:\Program Files\Tesseract-OCR\tesseract.exe")
TESSDATA_PREFIX = ROOT_DIR / "tessdata"
QWEN_ENDPOINT = "http://127.0.0.1:8910/v1/chat/completions"

if TESSERACT_CMD.exists():
    pytesseract.pytesseract.tesseract_cmd = str(TESSERACT_CMD)


@dataclass
class DocumentAsset:
    label: str
    name: str
    path: Optional[Path]
    mime: str
    size: int
    raw_bytes: bytes = b""
    page_count: int = 1
    text_excerpt: str = ""
    images: List[str] = field(default_factory=list)
    warnings: List[str] = field(default_factory=list)


def analyze_scene(req: SceneRequest, scene: str) -> Dict[str, Any]:
    max_pages = {"resume": 2, "contract": 3, "statement": 2, "paper": 10}.get(scene, 2)
    max_images = {"resume": 1, "contract": 0, "statement": 2, "paper": 3}.get(scene, 1)
    main_doc = load_document_asset(req.main_file, "main_file", max_pages=max_pages)
    supp_doc = None
    if req.supplement.file:
        supp_doc = load_document_asset(req.supplement.file, "supplement", max_pages=max_pages)

    prompt = build_scene_prompt(scene, req, main_doc, supp_doc)
    image_inputs: List[str] = []
    image_inputs.extend(main_doc.images[:max_images])
    if supp_doc:
        remaining = max(0, max_images - len(image_inputs))
        image_inputs.extend(supp_doc.images[:remaining])
    images_payload = image_inputs[:max_images]

    model_data: Dict[str, Any] = {}
    last_exc: Optional[Exception] = None
    try:
        model_data = call_qwen(prompt, images_payload)
    except Exception as exc:
        last_exc = exc
    if not model_data or _is_thin_model_data(model_data):
        reminder = "\n\nReminder: 上一次输出不合格（空/英文/格式错误）。请严格输出一行合法 JSON, 全部字符串字段用简体中文, 列表按 hint 给的下限填."
        try:
            retry_data = call_qwen(prompt + reminder, images_payload)
            if retry_data and (not model_data or not _is_thin_model_data(retry_data)):
                model_data = retry_data
                last_exc = None
        except Exception as exc:
            if not model_data:
                last_exc = exc
    if not model_data and last_exc is not None:
        raise last_exc
    return build_runtime_result(scene, req, model_data, main_doc, supp_doc)


def _is_thin_model_data(model_data: Dict[str, Any]) -> bool:
    if not isinstance(model_data, dict):
        return True
    facts = ensure_dict(model_data.get("facts"))
    judgement = ensure_dict(model_data.get("judgement"))
    summary = collapse_ws(judgement.get("summary") or model_data.get("summary") or "")
    if len(summary) < 30:
        return True
    facts_filled = sum(1 for value in facts.values() if _has_value(value))
    return facts_filled < 2


def load_document_asset(file_info: Dict[str, Any], label: str, max_pages: int) -> DocumentAsset:
    path_value = str(file_info.get("local_path") or "").strip()
    inline_bytes = b""
    path = None
    if path_value:
        path = Path(path_value)
        if not path.exists():
            raise FileNotFoundError(f"{label} file not found: {path}")
    else:
        inline_bytes = decode_inline_file_bytes(file_info, label)
        if not inline_bytes:
            raise ValueError(f"{label} local_path or inline file content is required")

    asset = DocumentAsset(
        label=label,
        name=str(file_info.get("name") or (path.name if path else f"{label}.bin")),
        path=path,
        mime=resolve_mime(file_info, path),
        size=int(file_info.get("size") or (path.stat().st_size if path else len(inline_bytes))),
        raw_bytes=inline_bytes,
    )
    suffix = resolve_suffix(asset)
    if suffix == ".pdf":
        return load_pdf_asset(asset, max_pages)
    if suffix in {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}:
        return load_image_asset(asset)
    if suffix in {".txt", ".md", ".json", ".csv"}:
        asset.text_excerpt = truncate_text(collapse_ws(read_text_asset(asset)), 6000)
        return asset
    if suffix == ".xlsx":
        asset.text_excerpt = load_xlsx_text(asset)
        return asset
    raise ValueError(f"{label} unsupported file type: {suffix or 'unknown'}")


def load_pdf_asset(asset: DocumentAsset, max_pages: int) -> DocumentAsset:
    if asset.raw_bytes:
        doc = fitz.open(stream=asset.raw_bytes, filetype="pdf")
    elif asset.path:
        doc = fitz.open(asset.path)
    else:
        raise ValueError(f"{asset.label} PDF source is unavailable")
    asset.page_count = len(doc)
    sections: List[str] = []
    try:
        for page_index in range(min(max_pages, len(doc))):
            page = doc[page_index]
            native_text = collapse_ws(page.get_text("text"))
            png_bytes = render_page_png(page)
            asset.images.append(to_data_url(png_bytes))
            ocr_text = ocr_png_bytes(png_bytes)
            merged_text = native_text if len(native_text) >= 120 else merge_text(native_text, ocr_text)
            if not merged_text:
                asset.warnings.append(f"{asset.label} page {page_index + 1} OCR text is sparse")
            per_page_limit = 1500 if max_pages >= 5 else 900
            sections.append(f"[page {page_index + 1}]\n{truncate_text(merged_text, per_page_limit)}")
    finally:
        doc.close()
    total_limit = 6000 if max_pages >= 5 else 2500
    asset.text_excerpt = truncate_text("\n\n".join(item for item in sections if item.strip()), total_limit)
    return asset


def load_image_asset(asset: DocumentAsset) -> DocumentAsset:
    image_source = io.BytesIO(asset.raw_bytes) if asset.raw_bytes else asset.path
    if image_source is None:
        raise ValueError(f"{asset.label} image source is unavailable")
    with Image.open(image_source) as image:
        image = image.convert("RGB")
        image.thumbnail((1600, 1600))
        buffer = io.BytesIO()
        image.save(buffer, format="PNG")
        png_bytes = buffer.getvalue()
    asset.images.append(to_data_url(png_bytes))
    asset.text_excerpt = truncate_text(ocr_png_bytes(png_bytes), 2500)
    if not asset.text_excerpt:
        asset.warnings.append(f"{asset.label} OCR text is sparse")
    return asset


def load_xlsx_text(asset: DocumentAsset) -> str:
    from openpyxl import load_workbook

    workbook_source = io.BytesIO(asset.raw_bytes) if asset.raw_bytes else asset.path
    if workbook_source is None:
        raise ValueError(f"{asset.label} spreadsheet source is unavailable")
    workbook = load_workbook(workbook_source, read_only=True, data_only=True)
    lines: List[str] = []
    try:
        for sheet in workbook.worksheets[:3]:
            lines.append(f"[sheet] {sheet.title}")
            for row in sheet.iter_rows(max_row=25, values_only=True):
                cells = [collapse_ws(cell) for cell in row if collapse_ws(cell)]
                if cells:
                    lines.append(" | ".join(cells[:8]))
    finally:
        workbook.close()
    return truncate_text("\n".join(lines), 6000)


def decode_inline_file_bytes(file_info: Dict[str, Any], label: str) -> bytes:
    data_url = str(file_info.get("data_url") or "").strip()
    if data_url:
        match = re.match(r"^data:([^;,]+)?;base64,(.*)$", data_url, flags=re.S)
        if not match:
            raise ValueError(f"{label} data_url is not base64 encoded")
        return decode_base64_bytes(match.group(2), label)

    for key in ("content_base64", "bytes_base64", "base64"):
        raw = file_info.get(key)
        if raw:
            return decode_base64_bytes(str(raw), label)
    return b""


def decode_base64_bytes(value: str, label: str) -> bytes:
    normalized = re.sub(r"\s+", "", value or "")
    if not normalized:
        return b""
    try:
        return base64.b64decode(normalized, validate=True)
    except Exception as exc:
        raise ValueError(f"{label} inline file content is not valid base64") from exc


def resolve_mime(file_info: Dict[str, Any], path: Optional[Path]) -> str:
    mime = str(file_info.get("mime") or file_info.get("type") or "").strip()
    if mime:
        return mime
    if path:
        guessed = mimetypes.guess_type(path.name)[0]
        if guessed:
            return guessed
    name = str(file_info.get("name") or "").strip()
    guessed = mimetypes.guess_type(name)[0] if name else None
    return guessed or "application/octet-stream"


def resolve_suffix(asset: DocumentAsset) -> str:
    suffix = Path(asset.name).suffix.lower()
    if suffix:
        return suffix
    mime_map = {
        "application/pdf": ".pdf",
        "image/png": ".png",
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/bmp": ".bmp",
        "image/tiff": ".tiff",
        "image/webp": ".webp",
        "text/plain": ".txt",
        "text/markdown": ".md",
        "application/json": ".json",
        "text/csv": ".csv",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
    }
    return mime_map.get(asset.mime.lower(), "")


def read_text_asset(asset: DocumentAsset) -> str:
    if asset.raw_bytes:
        return asset.raw_bytes.decode("utf-8", errors="ignore")
    if asset.path:
        return asset.path.read_text(encoding="utf-8", errors="ignore")
    raise ValueError(f"{asset.label} text source is unavailable")


def render_page_png(page: fitz.Page) -> bytes:
    pix = page.get_pixmap(matrix=fitz.Matrix(2.0, 2.0), alpha=False)
    return pix.tobytes("png")


def ocr_png_bytes(png_bytes: bytes) -> str:
    if TESSDATA_PREFIX.exists():
        import os

        os.environ["TESSDATA_PREFIX"] = str(TESSDATA_PREFIX)
    with Image.open(io.BytesIO(png_bytes)) as image:
        text = pytesseract.image_to_string(image, lang="chi_sim+eng")
    return collapse_ws(text)


def to_data_url(png_bytes: bytes) -> str:
    with Image.open(io.BytesIO(png_bytes)) as image:
        image = image.convert("RGB")
        image.thumbnail((1400, 1800))
        output = io.BytesIO()
        image.save(output, format="JPEG", quality=80, optimize=True)
    return "data:image/jpeg;base64," + base64.b64encode(output.getvalue()).decode("ascii")


def merge_text(native_text: str, ocr_text: str) -> str:
    if native_text and ocr_text:
        return truncate_text(native_text + "\n" + ocr_text, 2500)
    return native_text or ocr_text


def call_qwen(prompt: str, image_data_urls: List[str]) -> Dict[str, Any]:
    if image_data_urls:
        content: Any = [{"type": "text", "text": prompt}]
        for item in image_data_urls:
            content.append({"type": "image_url", "image_url": {"url": item}})
    else:
        # The local Genie-compatible endpoint accepts plain string content
        # reliably for text-only requests, while array-style content may fail.
        content = prompt

    payload = {
        "model": "qwen2.5vl3b",
        "messages": [{"role": "user", "content": content}],
        "temperature": 0.1,
        "max_tokens": 8192,
    }
    response = None
    for attempt in range(3):
        response = requests.post(QWEN_ENDPOINT, json=payload, timeout=240)
        if response.status_code != 429:
            break
        time.sleep(1.5 * (attempt + 1))
    if response is None:
        raise RuntimeError("qwen request was not sent")
    if not response.ok:
        detail = response.text.strip()
        if len(detail) > 400:
            detail = detail[:400]
        raise requests.HTTPError(
            f"{response.status_code} Client Error for url: {QWEN_ENDPOINT} :: {detail}",
            response=response,
        )
    data = response.json()
    raw = (((data.get("choices") or [{}])[0].get("message") or {}).get("content") or "").strip()
    if not raw:
        raise ValueError("qwen returned empty content")
    return parse_json_object(raw)


def parse_json_object(text: str) -> Dict[str, Any]:
    fenced = re.search(r"```(?:json)?\s*(\{.*\})\s*```", text, flags=re.S)
    candidate = fenced.group(1) if fenced else text
    start = candidate.find("{")
    end = candidate.rfind("}")
    if start < 0 or end <= start:
        raise ValueError(f"qwen did not return JSON: {truncate_text(text, 300)}")
    snippet = candidate[start : end + 1]
    try:
        return json.loads(snippet)
    except json.JSONDecodeError:
        return json.loads(sanitize_json_text(snippet))


def sanitize_json_text(text: str) -> str:
    def next_non_ws(index: int) -> str:
        pos = index + 1
        while pos < len(text) and text[pos] in " \t\r\n":
            pos += 1
        return text[pos] if pos < len(text) else ""

    out: List[str] = []
    in_string = False
    escaped = False
    for index, char in enumerate(text):
        if in_string:
            if escaped:
                out.append(char)
                escaped = False
                continue
            if char == "\\":
                out.append(char)
                escaped = True
                continue
            if char == '"':
                nxt = next_non_ws(index)
                if nxt and nxt not in {",", "}", "]", ":"}:
                    out.append('\\"')
                    continue
                out.append(char)
                in_string = False
                continue
            if char == "\n":
                out.append("\\n")
                continue
            if char == "\r":
                out.append("\\r")
                continue
            if char == "\t":
                out.append("\\t")
                continue
            out.append(char)
            continue
        out.append(char)
        if char == '"':
            in_string = True
    return "".join(out)


def build_scene_prompt(
    scene: str,
    req: SceneRequest,
    main_doc: DocumentAsset,
    supp_doc: Optional[DocumentAsset],
) -> str:
    parts = [
        "Analyze the documents and return JSON only.",
        f"Scene: {scene}",
        "",
        "Main document:",
        describe_asset(main_doc),
        "",
    ]
    if supp_doc:
        parts.extend(["Supplement document:", describe_asset(supp_doc), ""])
    elif req.supplement.form:
        parts.extend(
            [
                "Supplement form JSON:",
                json.dumps(req.supplement.form, ensure_ascii=False),
                "",
            ]
        )
    else:
        parts.extend(["Supplement: none", ""])
    parts.extend(
        [
            "Rules:",
            "- 所有字符串字段必须用简体中文输出，英文原文必须翻译。专有名词（人名/会议名/Transformer/BLEU 等）可保留英文。",
            "- 列表字段按 schema hint 给的下限填，不要给空数组；若完全无证据可填空字符串 / 0 / [].",
            "- 不输出 markdown / code fence，整段 JSON 一行返回，字符串内不要换行。",
            "- evidence.snippet ≤ 80 字。",
            "",
            scene_schema_instructions(scene),
        ]
    )
    return "\n".join(parts)


def describe_asset(asset: DocumentAsset) -> str:
    text_limit = 3000 if asset.page_count >= 5 else 800
    return "\n".join(
        [
            f"- name: {asset.name}",
            f"- mime: {asset.mime}",
            f"- pages: {asset.page_count}",
            f"- OCR/text excerpt: {sanitize_prompt_text(asset.text_excerpt, limit=text_limit) or '[empty]'}",
        ]
    )


def scene_schema_instructions(scene: str) -> str:
    if scene == "resume":
        return """Return one-line JSON, all values in 简体中文:
{"facts":{"candidate_name":"","target_role":"","years_of_experience":"","current_focus":"","education_top":{"degree":"","major":"","school":"","year":"","gpa":""},"core_strengths":[],"skills":[],"job_match_summary":""},"judgement":{"decision":"recommended|pending|reject","confidence":"low|medium|high","fit_level":"strong|medium|weak","summary":"","score":0,"next_step":"","focus_points":[]},"projects":[{"name":"","role":"","highlights":"","tech_stack":""}],"weak_phrases":[{"phrase":"","why":""}],"interview_questions":[],"interview_focus":[{"topic":"","why":""}],"evidence":[{"field":"","snippet":"","source":"main_file|supplement"}],"warnings":[]}
Hints: projects 2-4 项, weak_phrases 2-4 项（简历中含糊或夸大的表达 + 原因）, interview_focus 2-4 项。decision 仅用于初筛沟通建议，不要输出录取/淘汰等最终决定；recommended 表示建议进入下一轮沟通，pending 表示建议补充信息后复核，reject 表示当前匹配度偏弱。next_step 必须写成温和、可执行的中文动作建议。"""
    if scene == "contract":
        return """Return one-line JSON, all values in 简体中文:
{"facts":{"contract_type":"","counterparty":"","party_a":"","party_b":"","amount":"","effective_date":"","delivery_deadline":"","governing_law":""},"judgement":{"decision":"sign|revise|legal_review","confidence":"low|medium|high","risk_level":"low|medium|high","summary":"","recommendation":""},"clauses":[{"name":"","summary":"","importance":"low|medium|high"}],"change_summary":[{"label":"","value":"","risk_hint":""}],"timeline":[{"phase":"","date":"","party":"甲方|乙方|双方","label":"","note":""}],"risk_points":[],"negotiation_list":[],"evidence":[{"field":"","snippet":"","source":"main_file|supplement"}],"warnings":[]}
Hints: clauses 4-8 项。change_summary 仅输出归纳后的疑似修改点，不做逐条红线 diff，不要长段粘贴原文，2-5 项。timeline 3-6 项，仅输出履约节点摘要，不要声称这是严格准确的法律时间轴；每项必须标 party 是谁的动作。若缺少对方修改版合同，可将 change_summary 留空，但仍需基于主合同输出签署建议和风险点。"""
    if scene == "statement":
        return """Return one-line JSON, all values in 简体中文:
{"facts":{"account_name":"","statement_month":"","recognized_amount":"","ledger_amount":"","tax_base":"","document_type":""},"judgement":{"decision":"review|reviewed|partial","confidence":"low|medium|high","summary":"","mismatch_amount":"","tax_action":""},"receipt_entries":[{"date":"","item":"","amount":"","note":""}],"statement_entries":[{"date":"","item":"","amount":""}],"matched_pairs":[{"receipt":"","statement":"","note":""}],"mismatches":[{"field":"","recognized":"","ledger":"","reason":""}],"actions":[],"metrics":[{"label":"","value":""}],"evidence":[{"field":"","snippet":"","source":"main_file|supplement"}],"warnings":[]}
Hints: receipt_entries 把可见手写单据每行都列出（3-10 项）, statement_entries 仅 supplement 是银行流水时列出（3-10 项）, metrics 3-5 项核验指标。"""
    return """Return one-line JSON, all values in 简体中文（人名/会议英文名/数据集英文名可保留原文）:
{"facts":{"paper_title":"","year":"","domain":"","dataset":"","primary_metric":"","authors":"","venue":"","keywords":[]},"judgement":{"decision":"recommended|review|deep_read|skim|skip|pending","confidence":"low|medium|high","summary":"","relevance":"high|medium|low","next_step":"","fit_summary":""},"research_question":"","methods":[{"method":"","core_steps":"","fit":""}],"results":[{"label":"","value":0,"note":""}],"contributions":[],"reading_points":[],"reference_value":[{"angle":"","takeaway":""}],"evidence":[{"field":"","snippet":"","source":"main_file|supplement"}],"warnings":[]}
Hints: methods 2-5 项, results 3-6 项（value 尽量给数字）, reference_value 2-4 项（用户研究方向能借鉴的角度 + 收获）。研究问题 / 总结 / 贡献 / 阅读建议 必须全部用中文重写，禁止照搬英文。"""


def build_runtime_result(
    scene: str,
    req: SceneRequest,
    model_data: Dict[str, Any],
    main_doc: DocumentAsset,
    supp_doc: Optional[DocumentAsset],
) -> Dict[str, Any]:
    facts = ensure_dict(model_data.get("facts"))
    judgement = normalize_judgement(scene, ensure_dict(model_data.get("judgement")))
    evidence = normalize_evidence(model_data.get("evidence"), main_doc, supp_doc)
    warnings = normalize_strings(model_data.get("warnings"))
    warnings.extend(main_doc.warnings)
    if supp_doc:
        warnings.extend(supp_doc.warnings)
    missing_inputs = infer_missing_inputs(scene, req)
    if missing_inputs:
        warnings.extend(missing_input_warnings(scene))
    summary = collapse_ws(judgement.get("summary") or model_data.get("summary") or "")
    empty_output = is_empty_model_output(facts, judgement, summary, model_data)
    if empty_output:
        warnings.append("模型未输出有效结果，建议重试或切换更大模型。")
    return {
        "request_id": req.request_id,
        "scene": scene,
        "status": "ok",
        "confidence": choice(judgement.get("confidence"), {"low", "medium", "high"}, "medium"),
        "degraded": bool(missing_inputs) or empty_output,
        "decision": normalize_decision(scene, judgement.get("decision")),
        "summary": summary,
        "facts": facts,
        "judgement": judgement,
        "evidence": evidence,
        "cards": build_cards(scene, facts, judgement, model_data),
        "warnings": dedupe_strings(warnings),
        "missing_inputs": missing_inputs,
    }


def infer_missing_inputs(scene: str, req: SceneRequest) -> List[str]:
    if scene == "resume" and not collapse_ws((req.supplement.form or {}).get("jd_text")):
        return ["supplement.form.jd_text"]
    if scene in {"contract", "statement"} and not req.supplement.file:
        return ["supplement.file"]
    if scene == "paper":
        form = req.supplement.form or {}
        if not any(collapse_ws(form.get(key)) for key in ("direction", "current_topic", "focus", "purpose")):
            return ["supplement.form"]
    return []


def missing_input_warnings(scene: str) -> List[str]:
    if scene == "resume":
        return ["缺少岗位 JD，当前结论偏保守。"]
    if scene == "contract":
        return ["缺少对方修改版合同，当前无法输出完整红线对比。"]
    if scene == "statement":
        return ["缺少银行流水，当前只能基于主单据做初步核对。"]
    if scene == "paper":
        return ["缺少研究上下文，相关性判断会偏保守。"]
    return []


def build_cards(scene: str, facts: Dict[str, Any], judgement: Dict[str, Any], model_data: Dict[str, Any]) -> List[Dict[str, Any]]:
    if scene == "resume":
        education = ensure_dict(facts.get("education_top"))
        education_preview = collapse_ws(
            " · ".join(
                [
                    education.get("degree", ""),
                    education.get("major", ""),
                    education.get("school", ""),
                ]
            )
        )
        projects = [ensure_dict(p) for p in ensure_list(model_data.get("projects"))[:4]]
        weak_phrases = [ensure_dict(w) for w in ensure_list(model_data.get("weak_phrases"))[:4]]
        interview_questions = normalize_strings(model_data.get("interview_questions"))[:6]
        interview_focus = [ensure_dict(item) for item in ensure_list(model_data.get("interview_focus"))[:4]]
        screening_decision = normalize_decision("resume", judgement.get("decision"))
        return [
            {
                "key": "info",
                "title": "候选人信息",
                "preview": join_preview([facts.get("candidate_name"), facts.get("target_role"), education_preview]),
                "detail_mode": "kv",
                "detail_payload": {
                    "姓名": facts.get("candidate_name") or "",
                    "目标岗位": facts.get("target_role") or "",
                    "工作年限": facts.get("years_of_experience") or "",
                    "当前方向": facts.get("current_focus") or "",
                    "最高学历": education_preview or "",
                    "学历详情": education or {},
                    "核心优势": normalize_strings(facts.get("core_strengths"))[:6],
                    "技能栈": normalize_strings(facts.get("skills"))[:8],
                    "JD 匹配总结": facts.get("job_match_summary") or "",
                },
            },
            {
                "key": "judgement",
                "title": "综合评价",
                "preview": judgement.get("summary") or "",
                "detail_mode": "rich-text",
                "detail_payload": {
                    "summary": judgement.get("summary") or "",
                    "fit_level": judgement.get("fit_level") or "",
                    "score": judgement.get("score") or 0,
                    "focus_points": normalize_strings(judgement.get("focus_points"))[:6],
                    "projects": [
                        {
                            "name": item.get("name", ""),
                            "role": item.get("role", ""),
                            "highlights": item.get("highlights", ""),
                            "tech_stack": item.get("tech_stack", ""),
                        }
                        for item in projects
                    ],
                    "weak_phrases": [
                        {"phrase": item.get("phrase", ""), "why": item.get("why", "")}
                        for item in weak_phrases
                    ],
                },
            },
            {
                "key": "interview_questions",
                "title": "面试问题",
                "preview": interview_questions[0] if interview_questions else "",
                "detail_mode": "rich-text",
                "detail_payload": {
                    "questions": interview_questions,
                    "focus": [
                        {"topic": item.get("topic", ""), "why": item.get("why", "")}
                        for item in interview_focus
                    ],
                },
            },
            {
                "key": "screening_decision",
                "title": "初筛结论",
                "preview": judgement.get("next_step") or judgement.get("summary") or "",
                "detail_mode": "rich-text",
                "detail_payload": {
                    "decision": screening_decision,
                    "next_step": judgement.get("next_step") or "",
                    "summary": judgement.get("summary") or "",
                    "weak_phrases": [
                        {"phrase": item.get("phrase", ""), "why": item.get("why", "")}
                        for item in weak_phrases
                    ],
                    "warnings": normalize_strings(model_data.get("warnings")),
                },
            },
        ]
    if scene == "contract":
        change_summary = [ensure_dict(item) for item in ensure_list(model_data.get("change_summary"))[:6]]
        diffs = [ensure_dict(item) for item in ensure_list(model_data.get("clause_diffs"))[:6]]
        timeline_items = [ensure_dict(item) for item in ensure_list(model_data.get("timeline"))[:8]]
        clauses = [ensure_dict(item) for item in ensure_list(model_data.get("clauses"))[:8]]
        summarized_changes = change_summary or [
            {
                "label": item.get("clause") or "",
                "value": item.get("impact") or item.get("theirs") or "",
                "risk_hint": item.get("modified_by") or "",
            }
            for item in diffs
            if _has_value(item.get("clause")) or _has_value(item.get("impact")) or _has_value(item.get("theirs"))
        ][:6]
        return [
            {
                "key": "contract_info",
                "title": "关键信息",
                "preview": join_preview([facts.get("contract_type"), facts.get("counterparty"), facts.get("amount")]),
                "detail_mode": "kv",
                "detail_payload": {
                    "合同类型": facts.get("contract_type") or "",
                    "对手方": facts.get("counterparty") or "",
                    "甲方": facts.get("party_a") or "",
                    "乙方": facts.get("party_b") or "",
                    "金额": facts.get("amount") or "",
                    "生效日": facts.get("effective_date") or "",
                    "交付截止": facts.get("delivery_deadline") or "",
                    "管辖法律": facts.get("governing_law") or "",
                    "关键条款": [
                        {"name": item.get("name", ""), "summary": item.get("summary", ""), "importance": item.get("importance", "")}
                        for item in clauses
                    ],
                },
            },
            {
                "key": "signing_recommendation",
                "title": "签署建议",
                "preview": judgement.get("recommendation") or judgement.get("summary") or "",
                "detail_mode": "rich-text",
                "detail_payload": {
                    "recommendation": judgement.get("recommendation") or "",
                    "summary": judgement.get("summary") or "",
                    "risk_level": judgement.get("risk_level") or "",
                    "risk_points": normalize_strings(model_data.get("risk_points"))[:6],
                    "negotiation_list": normalize_strings(model_data.get("negotiation_list"))[:6],
                },
            },
            {
                "key": "redline_diff",
                "title": "变更摘要",
                "preview": render_first_change_preview(summarized_changes),
                "detail_mode": "rich-text",
                "detail_payload": {
                    "summary": "以下为模型归纳的疑似修改点，用于辅助复核，不作为逐条红线比对结果。",
                    "changes": [
                        {
                            "label": item.get("label", ""),
                            "value": item.get("value", ""),
                            "risk_hint": item.get("risk_hint", ""),
                        }
                        for item in summarized_changes
                    ],
                },
            },
            {
                "key": "timeline",
                "title": "履约节点摘要",
                "preview": join_preview([facts.get("effective_date"), facts.get("delivery_deadline")], sep=" → "),
                "detail_mode": "dual-timeline",
                "detail_payload": {
                    "party_a": facts.get("party_a") or "甲方",
                    "party_b": facts.get("party_b") or "乙方",
                    "summary": "以下为模型提取的履约节点摘要，适合快速浏览，正式排期仍需回看合同原文。",
                    "events": [
                        {
                            "phase": item.get("phase") or item.get("label") or "",
                            "date": item.get("date") or "",
                            "party": normalize_party(item.get("party")),
                            "label": item.get("label") or item.get("phase") or "",
                            "note": item.get("note") or "",
                        }
                        for item in timeline_items
                    ],
                },
            },
        ]
    if scene == "statement":
        receipt_entries = [ensure_dict(item) for item in ensure_list(model_data.get("receipt_entries"))[:10]]
        statement_entries = [ensure_dict(item) for item in ensure_list(model_data.get("statement_entries"))[:10]]
        matched_pairs = [ensure_dict(item) for item in ensure_list(model_data.get("matched_pairs"))[:6]]
        mismatches = [ensure_dict(item) for item in ensure_list(model_data.get("mismatches"))[:6]]
        metrics = [ensure_dict(item) for item in ensure_list(model_data.get("metrics"))[:8]]
        return [
            {
                "key": "ocr_result",
                "title": "手写单据识别结果",
                "preview": join_preview([facts.get("account_name"), facts.get("statement_month"), f"{len(receipt_entries)} 条"]),
                "detail_mode": "table",
                "detail_payload": {
                    "summary": {
                        "户名": facts.get("account_name") or "",
                        "账期": facts.get("statement_month") or "",
                        "确认金额": facts.get("recognized_amount") or "",
                        "账簿金额": facts.get("ledger_amount") or "",
                        "计税基数": facts.get("tax_base") or "",
                        "文档类型": facts.get("document_type") or "",
                    },
                    "columns": ["日期", "事项", "金额", "备注"],
                    "rows": [
                        [item.get("date", ""), item.get("item", ""), item.get("amount", ""), item.get("note", "")]
                        for item in receipt_entries
                    ],
                },
            },
            {
                "key": "comparison",
                "title": "对比结论",
                "preview": judgement.get("summary") or "",
                "detail_mode": "rich-text",
                "detail_payload": {
                    "summary": judgement.get("summary") or "",
                    "matched_pairs": [
                        {"receipt": item.get("receipt", ""), "statement": item.get("statement", ""), "note": item.get("note", "")}
                        for item in matched_pairs
                    ],
                    "mismatches": [
                        {
                            "field": item.get("field", ""),
                            "recognized": item.get("recognized", ""),
                            "ledger": item.get("ledger", ""),
                            "reason": item.get("reason", ""),
                        }
                        for item in mismatches
                    ],
                    "statement_entries": [
                        {"date": item.get("date", ""), "item": item.get("item", ""), "amount": item.get("amount", "")}
                        for item in statement_entries
                    ],
                },
            },
            {
                "key": "tax_advice",
                "title": "税务申报建议",
                "preview": judgement.get("tax_action") or "",
                "detail_mode": "rich-text",
                "detail_payload": {
                    "summary": judgement.get("summary") or "",
                    "tax_action": judgement.get("tax_action") or "",
                    "actions": normalize_strings(model_data.get("actions"))[:6],
                },
            },
            {
                "key": "metrics",
                "title": "验算指标",
                "preview": join_preview([facts.get("recognized_amount"), facts.get("ledger_amount")]),
                "detail_mode": "metric-grid",
                "detail_payload": {
                    "metrics": [
                        {"label": item.get("label", ""), "value": item.get("value", "")}
                        for item in metrics
                    ]
                },
            },
        ]
    methods = [ensure_dict(item) for item in ensure_list(model_data.get("methods"))[:6]]
    results = [ensure_dict(item) for item in ensure_list(model_data.get("results"))[:6]]
    contributions = normalize_strings(model_data.get("contributions"))[:6]
    reading_points = normalize_strings(model_data.get("reading_points"))[:6]
    reference_value = [ensure_dict(item) for item in ensure_list(model_data.get("reference_value"))[:4]]
    keywords = normalize_strings(facts.get("keywords"))[:6]
    return [
        {
            "key": "research_question",
            "title": "研究问题",
            "preview": truncate_text(model_data.get("research_question") or "", 80),
            "detail_mode": "rich-text",
            "detail_payload": {
                "question": model_data.get("research_question") or "",
                "summary": judgement.get("summary") or "",
                "paper_title": facts.get("paper_title") or "",
                "authors": facts.get("authors") or "",
                "venue": facts.get("venue") or "",
                "year": facts.get("year") or "",
                "domain": facts.get("domain") or "",
                "keywords": keywords,
            },
        },
        {
            "key": "methods",
            "title": "使用方法",
            "preview": (methods[0].get("method", "") if methods else ""),
            "detail_mode": "table",
            "detail_payload": {
                "columns": ["方法", "核心步骤", "适用性"],
                "rows": [
                    [item.get("method", ""), item.get("core_steps", ""), item.get("fit", "")]
                    for item in methods
                ],
            },
        },
        {
            "key": "results",
            "title": "实验结果",
            "preview": facts.get("primary_metric") or "",
            "detail_mode": "chart+text",
            "detail_payload": {
                "series": [
                    {"label": item.get("label", ""), "value": item.get("value", 0), "note": item.get("note", "")}
                    for item in results
                ],
                "summary": judgement.get("summary") or "",
                "contributions": contributions,
                "dataset": facts.get("dataset") or "",
                "primary_metric": facts.get("primary_metric") or "",
            },
        },
        {
            "key": "relevance",
            "title": "相关性结论",
            "preview": judgement.get("next_step") or judgement.get("summary") or "",
            "detail_mode": "rich-text",
            "detail_payload": {
                "summary": judgement.get("summary") or "",
                "fit_summary": judgement.get("fit_summary") or "",
                "relevance": judgement.get("relevance") or "",
                "next_step": judgement.get("next_step") or "",
                "reading_points": reading_points,
                "reference_value": [
                    {"angle": item.get("angle", ""), "takeaway": item.get("takeaway", "")}
                    for item in reference_value
                ],
            },
        },
    ]


def normalize_evidence(value: Any, main_doc: DocumentAsset, supp_doc: Optional[DocumentAsset]) -> List[Dict[str, Any]]:
    out = []
    for raw in ensure_list(value)[:8]:
        item = ensure_dict(raw)
        source = "supplement" if item.get("source") == "supplement" else "main_file"
        file_name = supp_doc.name if source == "supplement" and supp_doc else main_doc.name
        out.append(
            {
                "evidence_id": f"{source}-{item.get('field') or len(out) + 1}",
                "source": source,
                "field": collapse_ws(item.get("field") or ""),
                "snippet": truncate_text(item.get("snippet") or "", 180),
                "file_name": file_name,
            }
        )
    return out


def ensure_dict(value: Any) -> Dict[str, Any]:
    return dict(value) if isinstance(value, dict) else {}


def ensure_list(value: Any) -> List[Any]:
    return list(value) if isinstance(value, list) else []


_HINT_PLACEHOLDER = re.compile(r"^\s*\d+\s*[-–~]\s*\d+\s*项\s*$")


def normalize_strings(value: Any) -> List[str]:
    out: List[str] = []
    for item in ensure_list(value):
        text = collapse_ws(item)
        if not text:
            continue
        if _HINT_PLACEHOLDER.match(text):
            continue
        out.append(text)
    return out


def collapse_ws(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def sanitize_prompt_text(value: Any, limit: int = 1000) -> str:
    text = str(value or "")
    text = text.replace("×", "x").replace("“", '"').replace("”", '"').replace("’", "'")
    text = "".join(char if char >= " " or char in "\n\t" else " " for char in text)
    return truncate_text(collapse_ws(text), limit)


def truncate_text(value: Any, limit: int) -> str:
    text = str(value or "").strip()
    if len(text) <= limit:
        return text
    return text[: limit - 3].rstrip() + "..."


def dedupe_strings(items: List[str]) -> List[str]:
    seen = set()
    out = []
    for item in items:
        text = collapse_ws(item)
        if not text or text in seen:
            continue
        seen.add(text)
        out.append(text)
    return out


def choice(value: Any, allowed: set, default: str) -> str:
    text = collapse_ws(value).lower()
    return text if text in allowed else default


def normalize_party(value: Any) -> str:
    text = collapse_ws(value)
    if not text:
        return ""
    lowered = text.lower()
    if any(token in text for token in ("甲方", "party a", "party_a", "buyer", "client")):
        return "甲方"
    if any(token in text for token in ("乙方", "party b", "party_b", "seller", "supplier", "vendor", "contractor")):
        return "乙方"
    if any(token in text for token in ("双方", "both", "joint", "mutual")):
        return "双方"
    if lowered in {"a", "甲"}:
        return "甲方"
    if lowered in {"b", "乙"}:
        return "乙方"
    return ""


def normalize_decision(scene: str, value: Any) -> str:
    allowed = {
        "resume": {"recommended", "reject", "pending"},
        "contract": {"sign", "revise", "legal_review"},
        "statement": {"review", "reviewed", "partial"},
        "paper": {"recommended", "review", "deep_read", "skim", "skip", "pending"},
    }
    defaults = {"resume": "pending", "contract": "legal_review", "statement": "review", "paper": "review"}
    raw = collapse_ws(value).lower()
    if scene == "resume" and raw == "review":
        return "pending"
    if scene == "contract" and raw == "review":
        return "legal_review"
    return choice(value, allowed.get(scene, set()), defaults[scene])


JUDGEMENT_ENUM_FIELDS = {
    "resume": {
        "decision": ({"recommended", "reject", "pending"}, "pending"),
        "confidence": ({"low", "medium", "high"}, "medium"),
        "fit_level": ({"strong", "medium", "weak"}, "medium"),
    },
    "contract": {
        "decision": ({"sign", "revise", "legal_review"}, "legal_review"),
        "confidence": ({"low", "medium", "high"}, "medium"),
        "risk_level": ({"low", "medium", "high"}, "medium"),
    },
    "statement": {
        "decision": ({"review", "reviewed", "partial"}, "review"),
        "confidence": ({"low", "medium", "high"}, "medium"),
    },
    "paper": {
        "decision": ({"recommended", "review", "deep_read", "skim", "skip", "pending"}, "review"),
        "confidence": ({"low", "medium", "high"}, "medium"),
        "relevance": ({"high", "medium", "low"}, "medium"),
    },
}


def normalize_judgement(scene: str, judgement: Dict[str, Any]) -> Dict[str, Any]:
    spec = JUDGEMENT_ENUM_FIELDS.get(scene)
    if not spec:
        return judgement
    cleaned = dict(judgement)
    for field, (allowed, default) in spec.items():
        if field not in cleaned:
            continue
        raw = cleaned.get(field)
        normalized = choice(raw, allowed, "")
        if normalized:
            cleaned[field] = normalized
        else:
            cleaned[field] = "" if not collapse_ws(raw) else default
    return cleaned


def is_empty_model_output(
    facts: Dict[str, Any],
    judgement: Dict[str, Any],
    summary: str,
    model_data: Dict[str, Any],
) -> bool:
    if collapse_ws(summary):
        return False
    facts_filled = sum(1 for value in facts.values() if _has_value(value))
    judgement_filled = any(
        _has_value(value)
        for key, value in judgement.items()
        if key not in {"decision", "confidence", "risk_level", "relevance", "fit_level"}
    )
    extras_filled = any(
        _has_meaningful_extra(model_data.get(key))
        for key in (
            "clause_diffs", "timeline", "risk_points", "mismatches", "actions",
            "metrics", "methods", "results", "reading_points", "interview_questions",
        )
    )
    if judgement_filled or extras_filled:
        return False
    return facts_filled <= 1


def _has_meaningful_extra(value: Any) -> bool:
    if not isinstance(value, list):
        return _has_value(value)
    for item in value:
        if isinstance(item, dict):
            filled = sum(1 for v in item.values() if _has_value(v))
            if filled >= 2:
                return True
            continue
        if _has_value(item):
            return True
    return False


def _has_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(collapse_ws(value))
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, (list, tuple, set)):
        return any(_has_value(item) for item in value)
    if isinstance(value, dict):
        return any(_has_value(v) for v in value.values())
    return True


def join_preview(values: List[Any], sep: str = " · ") -> str:
    items = [collapse_ws(item) for item in values if collapse_ws(item)]
    return sep.join(items[:3])


def render_first_diff_preview(diffs: List[Any]) -> str:
    if not diffs:
        return ""
    item = ensure_dict(diffs[0])
    return join_preview([item.get("clause"), item.get("impact")])


def render_first_change_preview(changes: List[Any]) -> str:
    if not changes:
        return ""
    item = ensure_dict(changes[0])
    return join_preview([item.get("label"), item.get("value")])
